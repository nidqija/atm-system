import tkinter as tk
import openpyxl
import os

wb = openpyxl.load_workbook('raziq.xlsx')
sheet = wb.active


class GUI(tk.Tk):


      def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)   

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        

        self.frames = {}
        for F in (loginpage , welcomepage , deposit , view , cashIn , newDeposit , withdraw , bill , bill1 , bill2 , farewell):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

           
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("loginpage")

      def show_frame(self, page_name):
        frame = self.frames[page_name]
        frame.tkraise()


#======login=======#

class loginpage(tk.Frame):
    
    def __init__(self, parent , controller):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active

        for row in sheet.rows:
     
            tk.Frame.__init__(self , parent)
            self.controller = controller
            self.configure(background='blue')
            title = tk.Label(self , text='Welcome to XYZ Bank!' , font=('Helvetica' , 35 , 'bold' ,) , bg='blue' , fg='yellow')
            title.pack(padx= 20 , pady= 30)

            my_passkey = tk.IntVar()

            label = tk.Entry(self , text = 'Enter Account Number:' , textvariable=my_passkey , font=('Helvetica' , 25) , bg='white' )
            label.pack()
       
            button = tk.Button(self , text = 'Submit' , font=('Helvetica' , 20) ,height=1 , width= 30 , bg='yellow' ,command=lambda:check_password())
            button.pack(padx = 10 , pady= 20)

            incorrect_pass_label = tk.Label(self , text="" , fg="white" , bg='blue' , font=('Helvetica' , 17))
            incorrect_pass_label.pack(fill='both' , expand=True , padx=10 , pady=20)

            def check_password():
           
                if my_passkey.get() == row[1].value:
                     controller.show_frame('welcomepage')


                else:
                     incorrect_pass_label['text'] = 'Incorrect Password'


      
#===========welcome page==============#


class welcomepage(tk.Frame):  
         
     def __init__(self, parent , controller):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active
        for row in sheet.rows:

          tk.Frame.__init__(self , parent)
          self.configure(background='blue')
          self.controller = controller
          title = tk.Label(self , text=f'Hello {row[0].value}', font=('Helvetica' , 35 , 'bold' ,) , bg='blue' , fg='yellow' )
          title.pack(padx= 20 , pady= 30)
          button1 = tk.Button(self , text= 'Deposit' , fg='black' , bg='yellow' , height=1 , width=20 , font=('Helvetica' , 25 ) , command=lambda:controller.show_frame('deposit'))
          button1.pack(padx=20 , pady=10)
          button1 = tk.Button(self , text= 'Withdraw' , fg='black' , bg='yellow' , height=1 , width=20 , font=('Helvetica' , 25 ) , command=lambda:controller.show_frame('withdraw'))
          button1.pack(padx=20 , pady=10)
          button1 = tk.Button(self , text= 'Bills' , fg='black' , bg='yellow' , height=1 , width=20 , font=('Helvetica' , 25 ) , command=lambda:controller.show_frame('bill'))
          button1.pack(padx=20 , pady=10)
          button1 = tk.Button(self , text= 'View Account' , fg='black' , bg='yellow' , height=1 , width=20 , font=('Helvetica' , 25 ) , command=lambda:controller.show_frame('view'))
          button1.pack(padx=20 , pady=10)




#====================deposit=======================#


class deposit(tk.Frame):
    def __init__(self , parent , controller):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active

        for row in sheet.rows:
           tk.Frame.__init__(self , parent)
           self.controller = controller
           self.configure(bg='blue')
           title = tk.Label(self , text='Your Current Deposit Amount :' , font=('Helvetica' , 25 , 'bold') , fg='yellow' , bg='blue')
           title.pack(padx=10, pady=20)
           deposit_statement = tk.Label(self , text=f'RM {row[2].value}' , font=('Helvetica' , 40 , 'bold') , fg='yellow' , bg='blue')
           deposit_statement.pack(pady=10)
           pay = tk.Button(self , text='Cash In' , font=('Helvetica' , 25) , fg='black' , bg='yellow' , width=10 , command=lambda:controller.show_frame('cashIn'))
           pay.pack(padx=30 , pady=30)
           back = tk.Button(self , text='Return' , font=('Helvetica' , 25) , fg='black' , bg='yellow' , width=10 , command=lambda:controller.show_frame('welcomepage'))
           back.pack(padx=10 , pady=20)





#================view statement=================#       



class view(tk.Frame):
    def __init__(self, parent , controller ):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active
        wb.save('raziq.xlsx')

        for row in sheet.rows:

           tk.Frame.__init__(self , parent)
           self.controller = controller
           self.configure(bg='blue')

           
           label = tk.Label(self , text='Below are your current statement:' , fg='yellow' , bg='blue' , font=('Helvetica' , 25))
           label.pack(padx=10 , pady=10)
           label = tk.Label(self , text=f'Account Name : {row[0].value}' , fg='yellow' , bg='blue' , font=('Helvetica' , 25 , 'bold') , justify='left' , anchor='w')
           label.pack(padx=10 , pady=2)
           label = tk.Label(self , text=f' Balance : RM{row[2].value}' , fg='yellow' , bg='blue' , font=('Helvetica' , 25 , 'bold') , justify='left' , anchor='w')
           label.pack(padx=10 , pady=2)
           label = tk.Label(self , text=f' TNB Bill : RM{row[3].value}' , fg='yellow' , bg='blue' , font=('Helvetica' , 25 , 'bold'))
           label.pack(padx=10 , pady=2)
           label = tk.Label(self , text=f' Water Bill : RM{row[4].value}' , fg='yellow' , bg='blue' , font=('Helvetica' , 25 , 'bold'))
           label.pack(padx=10 , pady=2)
           label = tk.Label(self , text=f'ASTRO Bill : RM{row[5].value}' , fg='yellow' , bg='blue' , font=('Helvetica' , 25 , 'bold'))
           label.pack(padx=10 , pady=2)
           back = tk.Button(self , text='Return' , font=('Helvetica' , 20) , fg='black' , bg='yellow' , width=10 , command=lambda:controller.show_frame('welcomepage'))
           back.pack(padx=10 , pady=10)
           finish = tk.Button(self , text='Finish' , fg='black' , bg='yellow' , font=('Helvetica' , 20) , width=10 , command=lambda:controller.show_frame('farewell'))
           finish.pack(padx=10 , pady=20)
            
            
           
#=============== cash in(deposit)======================#


class cashIn(tk.Frame):
    def __init__(self, parent , controller):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active

        for row in sheet.rows:
           tk.Frame.__init__(self , parent)
           self.controller= controller
           self.configure(bg='blue')
           rusure = tk.Label(self , text='Enter Amount:' , fg='yellow' , bg='blue' , font=('Helvetica' , 35 , 'bold'))
           rusure.pack(pady=10 , padx=40)
           entry_deposit = tk.IntVar()
           entry2 = tk.Entry(self , font=('Helvetica' , 35 ) , textvariable=entry_deposit)
           entry2.pack(padx=10 , pady=20)
           button = tk.Button(self , text='Submit' , fg='black' , bg='yellow' , font=('Helvetica' , 25) , command=lambda:checkDepo())
           button.pack()
           invalid = tk.Label(self , text='' , fg='yellow' , bg='blue' , font=('Helvetica' , 25))
           invalid.pack(padx=10 , pady=10)
        


           def checkDepo():
              try:
                    float(entry2.get())
                    rusure['text'] = 'Are you sure?'
                    button2 = tk.Button(self , text='Yes'  ,bg='yellow' , fg='black' , font=('Helvetica' , 25) , width=20 , command=lambda:sum())
                    button2.pack(pady=5)
                    button3 = tk.Button( self , text='No' , bg='yellow' , fg='black' , font=('Helvetica' , 25) , width=20 , command=lambda:no())
                    button3.pack()
                    button.forget()

                    def no():
                            rusure['text'] = 'Enter Amount :'
                            button2.forget()
                            button3.forget()
                            button.pack()

                    def sum():   
                       

                           new_money =  float(entry2.get()) + float(row[2].value)
                           row[2].value = new_money
                           rusure['text'] = f' Your new balance  {row[2].value}'
                           entry2.forget()
                           button2.forget()
                           button3.forget()
                           comeback = tk.Button(self , text='Return' , bg='yellow' , fg='black' , font=('Helvetica' , 25) , command=lambda:restart() )
                           comeback.pack()
                           wb.save('raziq.xlsx')

                    def restart():
                       app.destroy()
                       os.startfile('ui.pyw')
              
                    wb.save('raziq.xlsx')

              except ValueError:
                    invalid['text'] = 'Invalid Input'


        wb.save('raziq.xlsx')






#--------withdraw----------------------------#


class withdraw(tk.Frame):
     def __init__(self, parent , controller):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active

        for row in sheet.rows:
         
            tk.Frame.__init__(self ,parent)
            self.controller = controller
            self.configure(bg='blue')  
            current_depo = tk.Label(self , text=f'Your current balance : RM {row[2].value}' , fg='yellow' , bg='blue' , font=('Helvetica' , 25 , 'bold'))
            current_depo.pack(padx=10 , pady= 40)
            amount = tk.Label(self , text='Enter Amount:' , fg='yellow' , bg='blue' , font=('Helvetica' , 20 , 'bold'))
            amount.pack(pady=10 , padx=40)
            entry_deposit = tk.IntVar()
            entry2 = tk.Entry(self , font=('Helvetica' , 35 ) , textvariable=entry_deposit)
            entry2.pack(padx=10 , pady=20)
            button = tk.Button(self , text='Submit' , fg='black' , bg='yellow' , font=('Helvetica' , 25) , command=lambda:checkAmount())
            button.pack()
            goback = tk.Button( self , text='Return' , fg='black' , bg='yellow' , font=('Helvetica' , 25) , command=lambda:controller.show_frame('welcomepage'))
            goback.pack(padx=10 , pady=20)
            invalid = tk.Label(self , text='' , fg='yellow' , bg='blue' , font=('Helvetica' , 25))
            invalid.pack(padx=10 , pady=10)


            def checkAmount():

               try:
                  float(entry2.get())
                  current_depo['text'] = 'Are you sure?'
                  agree = tk.Button(self , text='Yes' , fg='black' , bg='yellow' , font=('Helvetica' , 25) , command=lambda:substract())
                  agree.pack()
                  button.forget()
                  goback.forget()


                  def substract():
                      new_money =  float(row[2].value) - float(entry2.get())
                      row[2].value = new_money
                      current_depo['text'] = f'Thank you!'
                      amount['text'] = f'Your current balance is : RM {row[2].value}'
                      comeback = tk.Button(self , text='Return' , bg='yellow' , fg='black' , font=('Helvetica' , 25) , command=lambda:restart() )
                      comeback.pack()
                      entry2.forget()
                      agree.forget()
                      wb.save('raziq.xlsx')


                  def restart():
                      app.destroy()
                      os.startfile('ui.pyw')


                  
                     






               except ValueError:
                  invalid['text'] = 'invalid input'
                  
                  

#========================newdepo (withdraw)======================#



class newDeposit(tk.Frame):
    def __init__(self, parent , controller):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active
        wb.save('raziq.xlsx')

        

        for row in sheet.rows:
          
          tk.Frame.__init__(self , parent)
          self.controller = controller
          self.configure(bg='blue')
          congrats = tk.Label(self , text= ' Thank you for your service!' , fg='yellow' , bg='blue' , font=('Helvetica' , 35) )
          congrats.pack(padx=10 , pady=20)
          label = tk.Label(self , text= f'Your new balance : RM {row[2].value}', fg='yellow' , bg='blue' , font=('Helvetica' , 30 , 'bold'))
          label.pack(padx=10 , pady=20)
          resume = tk.Button(self , text='Return' , fg='black' , bg='yellow' , font=('Helvetica' , 25) , width=10 , command=lambda:controller.show_frame('welcomepage'))
          resume.pack(padx=10 , pady=20)
          finish = tk.Button(self , text='Finish' , fg='black' , bg='yellow' , font=('Helvetica' , 25) , width=10 , command=lambda:controller.show_frame('farewell'))
          finish.pack(padx=10 , pady=20)





#=============================bill======================================#

class bill(tk.Frame):
    def __init__(self, parent , controller):
        tk.Frame.__init__(self , parent)
        self.controller = controller
        self.configure(bg='blue')
        title = tk.Label(self , text='Here are your bills :' , fg='yellow' , font=('Helvetica' , 35  , 'bold') , bg='blue')
        title.pack(padx=10 , pady=30)
        bill1 = tk.Button(self , text='TNB bill' , bg='blue' , font=('Helvetica' , 25  , 'bold') , fg='yellow' , width=30 , command=lambda:controller.show_frame('bill1'))
        bill1.pack(padx=10 , pady=10)
        bill2 = tk.Button(self , text='Water bill' , bg='blue' , font=('Helvetica' , 25  , 'bold') , fg='yellow' , width=30 , command=lambda:controller.show_frame('bill2'))
        bill2.pack(padx=10 , pady=10)
        bill3 = tk.Button(self , text='ASTRO bill' , bg='blue' , font=('Helvetica' , 25  , 'bold') , fg='yellow' , width=30 )
        bill3.pack(padx=10 , pady=10)  
        comeback = tk.Button(self , text='Return' , bg='yellow' , fg='black' , font=('Helvetica' , 25) , command=lambda:controller.show_frame('welcomepage'))
        comeback.pack(padx=10 , pady=10) 





#===============================bill 1 ( tnb bill)=====================================#

class bill1(tk.Frame):
    def __init__(self , parent , controller):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active

        for row in sheet.rows:
           tk.Frame.__init__(self ,parent)
           self.controller = controller
           self.configure(bg='blue')
           bill1 = tk.Label(self , text='Your TNB bill is :' ,  fg='yellow' , font=('Helvetica' , 35  , 'bold') , bg='blue')
           bill1.pack(padx=10 , pady=10)
           price = tk.Label(self , text=f'RM {row[3].value}' , font=('Helvetica' , 40  , 'bold') , fg='yellow' , bg='blue' )
           price.pack(padx=10 , pady=20)
           bill1_deposit = tk.IntVar()
           entry3 = tk.Entry(self , textvariable= bill1_deposit , width=25 , font=('Helvetica' , 30))
           entry3.pack()
           error = tk.Label(self , text='' , fg='yellow' , bg='blue' ,  font=('Helvetica' , 25) )
           error.pack()
           pay = tk.Button(self , text='Pay' , width=10 , fg='black' , bg='yellow' , font=('Helvetica' , 25) , command=lambda:deducbill1())
           pay.pack(pady=20 , padx=10)
           comeback = tk.Button(self , text='Return' , width=10 , fg='black' , bg='yellow' , font=('Helvetica' , 25) , command=lambda:controller.show_frame('bill'))
           comeback.pack(pady=20 , padx=10)

           def deducbill1():
               try:
                float(entry3.get())
                price.forget()
                bill1['text'] = 'Are you sure?'
                pay.forget()
                comeback.forget()
                agree = tk.Button(self , text='yes' , fg='black' , bg='yellow' , font=('Helvetica' , 25) , width=20 , command=lambda:deduc2()  )
                agree.pack(padx=10 , pady=20)
                stepback = tk.Button(self , text='no' ,  fg='black' , bg='yellow' , font=('Helvetica' , 25) , width=20)
                stepback.pack()
                



               except ValueError:
                   error['text'] = 'Invalid input!'




               def deduc2():
                      minus = float(row[3].value) - float(entry3.get())
                      row[3].value = minus
                      bill1['text'] = 'Thank You!'
                      price2 = tk.Label(self , text=f'Your current balance is RM {row[3].value}' , bg='blue' , fg='yellow' , font=('Helvetica' , 30 , 'bold') ) 
                      price2.pack(padx=10 , pady=20)
                      entry3.forget()
                      agree.forget()
                      stepback.forget()
                      comeback = tk.Button(self , text='Return' , bg='yellow' , fg='black' , font=('Helvetica' , 25) , command=lambda:restart() )
                      comeback.pack(padx=10  ,pady=10)
                      wb.save('raziq.xlsx')


               def restart():
                    app.destroy()
                    os.startfile('ui.pyw')

                   
                   
                   








#===============================bill 2 ( water bill)===========================#
           
class bill2(tk.Frame):
    def __init__(self, parent , controller):
        wb = openpyxl.load_workbook('raziq.xlsx')
        sheet = wb.active

        for row in sheet.rows:
           tk.Frame.__init__(self , parent)
           self.controller = controller
           self.configure(bg='blue')
           bill2 = tk.Label(self , text='Your Water bill is :' ,  fg='yellow' , font=('Helvetica' , 35  , 'bold') , bg='blue' )
           bill2.pack()
           price = tk.Label(self , text=f'RM {row[4].value}' , font=('Helvetica' , 40  , 'bold') , fg='yellow' , bg='blue' )
           price.pack(padx=10 , pady=20)
           entry3 = tk.Entry(self , width=25 , font=('Helvetica' , 30))
           entry3.pack()
           pay = tk.Button(self , text='Pay' , width=10 , fg='black' , bg='yellow' , font=('Helvetica' , 25))
           pay.pack(pady=20 , padx=10)
           comeback = tk.Button(self , text='Return' , width=10 , fg='black' , bg='yellow' , font=('Helvetica' , 25) , command=lambda:controller.show_frame('bill'))
           comeback.pack(pady=20 , padx=10)


#================================end program==========================================#
    

class farewell(tk.Frame):
    def __init__(self ,parent , controller):
           tk.Frame.__init__(self , parent)
           self.controller= controller
           self.configure(bg='blue')
           label = tk.Label( self , text='Thank you!' , fg='yellow' , bg='blue' , font=('Helvetica' , 35 , 'bold') )
           label.pack(padx=10 , pady=30)
           label2 = tk.Label( self , text='We hope to see you again!' , fg='yellow' , bg='blue' , font=('Helvetica' , 35 , 'bold'))
           label2.pack(padx=10 , pady=30)
           



        

           


        
        

           

           
               
               




          
                   
           










if __name__ == "__main__":
      app = GUI()
      app.geometry('700x500')
      app.mainloop()
