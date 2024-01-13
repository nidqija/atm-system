import tba.atm_data as atm_data
import openpyxl 



wb = openpyxl.load_workbook("raziq.xlsx")
ws = wb['Sheet1']


methods = ( "1. Deposit" , "2.Withdraw" , "3.Balance")
lines = print("----------------------")
bills = ("1. TNB" , "2. Air" , "3. Astro")





print("Welcome to Maybank!")

while True:
  passkey = int(input("Enter your passkey:"))
  

 
  if passkey == (ws['B2'].value):
      print( "Welcome!" , (ws['A2'].value))
  
  for method in methods:
   print(lines)
   print(method)

  question = input("which one?")

  if question == "1":
     balance = print(f"Your balance is now {ws['C2'].value}")
     deposit = float(input("How much amount do you want to add?: "))
     b_col = ws['C2'].value
     new_balance =  deposit + b_col
     print(f"Your new balance: RM{new_balance}")
     ws['C2'].value = new_balance
     wb.save('raziq.xlsx')
      
     
   

  elif question == "2":
     amount = print("Your current balance is:" , (ws['C2']).value)
     withdraw = float(input("How much do you want to withdraw?:"))
     b_col = ws['C2'].value
     new_amount = b_col - withdraw
     print(f"Your new amount is :RM{new_amount}")
     ws['C2'].value = new_amount
     wb.save('raziq.xlsx')
    


  elif question == "3":

     for bill in bills:
          print(lines)
          print(bill)


          

     current_bill = str(input("Enter Bill type:"))

     if current_bill == '1':
          print(f'your tnb bill is RM' , (ws['D2']).value)
          pay = float(input('Enter payment:'))
          d_col = (ws['D2']).value
          deduc = d_col - pay
          (ws['D2']).value = deduc
          wb.save('raziq.xlsx')
          
          if pay < deduc : 
               print(f'you have RM{deduc} left to pay ')
               rewind = str(input('do you want to pay more?(y/n)'))
               if rewind == 'n':
                     print(f'you current bill: RM{deduc}')
                     print('Thank you for using our service!')
               else:
                     pay = float(input('Enter payment:'))
                     d_col = (ws['D2']).value
                     deduc_plus1 = d_col - pay
                     deduc_plus1 = (ws['D2']).value
                     print(f'you have RM{deduc_plus1} left to pay')
                     print('Thank you for using our service!')
                     wb.save('raziq.xlsx')
                     

                    
                    
               
                    
                  

          else:
               print(f'you have RM{deduc} left to pay')
               print(f'Congratulations! You are fully-paid')
               print('Thank you for using our service!')
                    
                          
                                              
            





     elif current_bill == '2':
          print(f'your water bill is RM{atm_data.air}')
          pay = float(input('Enter payment:'))
          deduc_2 = atm_data.air - pay
          
          if pay < atm_data.air : 
               print(f'you have RM{deduc_2} left to pay ')
               rewind = str(input('do you want to pay more?(y/n)'))
               if rewind == 'n':
                  print(f'you current bill: RM{deduc_2}')
                  print('Thank you for using our service!')
               else:
                     pay = float(input('Enter payment:'))
                     deduc_2plus1 = deduc_2 - pay
                     print(f'you have RM{deduc_2plus1} left to pay')
                     print('Thank you for using our service!')
                     

                    
                    
               
                    
                  

          else:
               print(f'you have RM{deduc_2} left to pay')
               print(f'Congratulations! You are fully-paid')
               print('Thank you for using our service!')



     elif current_bill == '3':
          print(f'your tv broadcast bill is RM{atm_data.astro}')
          pay = float(input('Enter payment:'))
          deduc_3 = atm_data.astro - pay
          
          if pay < atm_data.astro : 
               print(f'you have RM{deduc_3} left to pay ')
               rewind = str(input('do you want to pay more?(y/n)'))
               if rewind == 'n':
                  print(f'you current bill: RM{deduc_3}')
                  print('Thank you for using our service!')
               else:
                     pay = float(input('Enter payment:'))
                     deduc_3plus1 = deduc_3 - pay
                     print(f'you have RM{deduc_3plus1} left to pay')
                     print('Thank you for using our service!')
               
          else:
               print(f'you have RM{deduc_3} left to pay')
               print(f'Congratulations! You are fully-paid')


     

     
  if input('Do you want to restart? (y/n):') ==  'n':
     print("Thank you for using our service!")
     break
  



wb.save("raziq.xlsx")

                    









              




      

          
     
     


     
     

   







                    









              




      

          
     
     


     
     

   







                    









              




      

          
     
     


     
     

   











              




      

          
     
     


     
     

   





